import codecs
import csv
from io import BytesIO, StringIO

from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import load_workbook
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework import status
from rest_framework.test import APITestCase

from apps.accounts.models import UserProfile
from apps.assignments.models import Assignment
from apps.audit.models import AuditLog
from apps.employees.models import Department, Employee, EmployeeImportJob, JobTitle
from apps.employees.views import repair_mojibake_for_export
from apps.inventory.models import Asset, AssetCategory
from apps.tickets.models import Ticket

User = get_user_model()

REAL_WORLD_MOJIBAKE_DEPARTMENT = (
    "\u00c3\u201e\u00c2\u00b0DAR\u00c3\u201e\u00c2\u00b0 VE "
    "MAL\u00c3\u201e\u00c2\u00b0 \u00c3\u201e\u00c2\u00b0"
    "\u00c3\u2026\u00c2\u017eLER M\u00c3\u0192\u00c5\u201cD"
    "\u00c3\u0192\u00c5\u201cRL\u00c3\u0192\u00c5\u201c"
    "\u00c3\u201e\u00c2\u017e\u00c3\u0192\u00c5\u201c"
)
REPAIRED_REAL_WORLD_DEPARTMENT = (
    "\u0130DAR\u0130 VE MAL\u0130 \u0130\u015eLER "
    "M\u00dcD\u00dcRL\u00dc\u011e\u00dc"
)
MOJIBAKE_DEPARTMENT = (
    "\u00c3\u201e\u00c2\u00b0\u00c3\u2026\u00c2\u017eLETMELER "
    "M\u00c3\u0192\u00c5\u201cD\u00c3\u0192\u00c5\u201cRL"
    "\u00c3\u0192\u00c5\u201c\u00c3\u201e\u00c2\u017e\u00c3\u0192\u00c5\u201c"
)
REPAIRED_DEPARTMENT = "\u0130\u015eLETMELER M\u00dcD\u00dcRL\u00dc\u011e\u00dc"
MOJIBAKE_JOB_TITLE = (
    "\u00c3\u201e\u00c2\u00b0dari ve Mali "
    "\u00c3\u201e\u00c2\u00b0\u00c3\u2026\u00c5\u00b8ler "
    "M\u00c3\u0192\u00c2\u00bcd\u00c3\u0192\u00c2\u00bcr"
    "\u00c3\u0192\u00c2\u00bc"
)
REPAIRED_JOB_TITLE = "\u0130dari ve Mali \u0130\u015fler M\u00fcd\u00fcr\u00fc"


class EmployeeApiTests(APITestCase):
    def create_user_with_role(self, username, role):
        user = User.objects.create_user(
            username=username,
            email=f"{username}@example.com",
            password="StrongPass123!",
        )

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.role = role
        profile.save(update_fields=["role"])

        return User.objects.get(pk=user.pk)

    def setUp(self):
        self.admin_user = self.create_user_with_role(
            "admin-user",
            UserProfile.Role.ADMIN,
        )
        self.technician_user = self.create_user_with_role(
            "technician-user",
            UserProfile.Role.TECHNICIAN,
        )
        self.viewer_user = self.create_user_with_role(
            "viewer-user",
            UserProfile.Role.VIEWER,
        )
        self.requester_user = self.create_user_with_role(
            "requester-user",
            UserProfile.Role.REQUESTER,
        )

        self.department = Department.objects.create(name="Bilgi İşlem")
        self.job_title = JobTitle.objects.create(name="Uzman")
        self.asset_category = AssetCategory.objects.create(name="Laptop")

        self.manager_user = self.create_user_with_role(
            "manager-user",
            UserProfile.Role.APPROVER,
        )
        self.manager = Employee.objects.create(
            user=self.manager_user,
            full_name="Manager Personel",
            email="manager@example.com",
            department=self.department,
            job_title=self.job_title,
            is_active=True,
        )

        self.employee = Employee.objects.create(
            user=self.requester_user,
            manager=self.manager,
            full_name="Requester Personel",
            employee_code="EMP-001",
            email="requester@example.com",
            phone="5551112233",
            department=self.department,
            job_title=self.job_title,
            external_hr_id="HR-001",
            sync_source=Employee.SyncSource.MANUAL,
            is_active=True,
        )

    def create_asset(self, *, name, inventory_code):
        return Asset.objects.create(
            category=self.asset_category,
            name=name,
            inventory_code=inventory_code,
            serial_number=f"SN-{inventory_code}",
            status=Asset.Status.ACTIVE,
        )

    def test_legacy_employee_endpoint_still_returns_array(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)

    def test_employee_table_endpoint_returns_paginated_response(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/table/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("count", response.data)
        self.assertIn("next", response.data)
        self.assertIn("previous", response.data)
        self.assertIn("results", response.data)

    def test_employee_table_supports_search_and_user_role_filter(self):
        Employee.objects.create(
            full_name="Bağımsız Personel",
            email="bagimsiz@example.com",
            department=self.department,
            job_title=self.job_title,
            is_active=True,
        )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get(
            "/api/employees/table/",
            {
                "search": "Requester",
                "user_role": UserProfile.Role.REQUESTER,
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["id"], self.employee.id)

    def test_employee_detail_returns_profile_user_org_assignments_and_tickets(self):
        active_asset = self.create_asset(
            name="Aktif Zimmet Laptop",
            inventory_code="IT-LPT-001",
        )
        returned_asset = self.create_asset(
            name="İade Edilmiş Laptop",
            inventory_code="IT-LPT-002",
        )

        active_assignment = Assignment.objects.create(
            asset=active_asset,
            employee=self.employee,
            assigned_by=self.admin_user,
            notes="Güncel zimmet",
        )
        Assignment.objects.create(
            asset=returned_asset,
            employee=self.employee,
            assigned_at=timezone.now() - timezone.timedelta(days=10),
            returned_at=timezone.now() - timezone.timedelta(days=1),
            assigned_by=self.admin_user,
            returned_by=self.admin_user,
        )

        Ticket.objects.create(
            employee=self.employee,
            title="Açık ticket",
            description="Açık ticket açıklaması",
            category=Ticket.Category.ACCESS,
            priority=Ticket.Priority.HIGH,
            status=Ticket.Status.OPEN,
            created_by=self.requester_user,
        )
        Ticket.objects.create(
            employee=self.employee,
            title="İşlemde ticket",
            description="İşlemde ticket açıklaması",
            category=Ticket.Category.HARDWARE,
            priority=Ticket.Priority.NORMAL,
            status=Ticket.Status.IN_PROGRESS,
            created_by=self.requester_user,
        )
        Ticket.objects.create(
            employee=self.employee,
            title="Çözüldü ticket",
            description="Çözüldü ticket açıklaması",
            category=Ticket.Category.SOFTWARE,
            priority=Ticket.Priority.NORMAL,
            status=Ticket.Status.RESOLVED,
            created_by=self.requester_user,
        )
        Ticket.objects.create(
            employee=self.employee,
            title="Kapandı ticket",
            description="Kapandı ticket açıklaması",
            category=Ticket.Category.OTHER,
            priority=Ticket.Priority.LOW,
            status=Ticket.Status.CLOSED,
            created_by=self.requester_user,
        )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get(f"/api/employees/{self.employee.id}/detail/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(response.data["employee"]["id"], self.employee.id)
        self.assertEqual(response.data["employee"]["manager"]["id"], self.manager.id)
        self.assertEqual(response.data["employee"]["department"]["name"], "Bilgi İşlem")
        self.assertEqual(response.data["employee"]["job_title"]["name"], "Uzman")

        self.assertEqual(response.data["user"]["username"], self.requester_user.username)
        self.assertEqual(response.data["user"]["role"], UserProfile.Role.REQUESTER)

        self.assertEqual(response.data["summary"]["active_assignment_count"], 1)
        self.assertEqual(response.data["summary"]["total_assignment_count"], 2)
        self.assertEqual(response.data["summary"]["open_ticket_count"], 1)
        self.assertEqual(response.data["summary"]["in_progress_ticket_count"], 1)
        self.assertEqual(response.data["summary"]["resolved_ticket_count"], 1)
        self.assertEqual(response.data["summary"]["closed_ticket_count"], 1)
        self.assertEqual(response.data["summary"]["total_ticket_count"], 4)

        active_assignment_ids = {
            item["id"] for item in response.data["active_assignments"]
        }

        self.assertEqual(active_assignment_ids, {active_assignment.id})
        self.assertEqual(len(response.data["recent_tickets"]), 4)

    def test_viewer_can_read_employee_detail(self):
        self.client.force_authenticate(user=self.viewer_user)

        response = self.client.get(f"/api/employees/{self.employee.id}/detail/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_requester_cannot_read_employee_detail(self):
        self.client.force_authenticate(user=self.requester_user)

        response = self.client.get(f"/api/employees/{self.employee.id}/detail/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_can_export_employees_as_csv(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("attachment;", response["Content-Disposition"])

        content = response.content.decode("utf-8-sig")

        self.assertIn("Ad Soyad", content)
        self.assertIn("Requester Personel", content)
        self.assertIn("EMP-001", content)

    def export_csv_rows(self, response):
        content = response.content.decode("utf-8-sig")
        sep_line, csv_content = content.split("\n", 1)

        return sep_line.rstrip("\r"), list(
            csv.reader(StringIO(csv_content), delimiter=";"),
        )

    def exported_employee_row(self, response, employee):
        _, rows = self.export_csv_rows(response)

        return next(row for row in rows if row[0] == str(employee.id))

    def export_xlsx_workbook(self, response):
        return load_workbook(BytesIO(response.content))

    def create_import_upload(self, rows, filename="personnel-import.xlsx"):
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        for row in rows:
            worksheet.append(row)

        stream = BytesIO()
        workbook.save(stream)

        return SimpleUploadedFile(
            filename,
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def create_csv_import_upload(self, rows, filename="personnel-import.csv"):
        stream = StringIO()
        writer = csv.writer(stream, delimiter=";")
        writer.writerows(rows)

        return SimpleUploadedFile(
            filename,
            ("\ufeff" + stream.getvalue()).encode("utf-8"),
            content_type="text/csv",
        )

    def exported_employee_xlsx_row(self, response, employee):
        worksheet = self.export_xlsx_workbook(response)["Personel"]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] == employee.id:
                return row

        self.fail(f"Employee {employee.id} row not found in XLSX export")

    def test_employee_import_dry_run_xlsx_returns_preview_without_db_write(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "E-posta", "Departman", "Unvan"],
                ["Çağrı Şahin", "IMP-001", "cagri.import@example.com", "Yeni Departman", "Yeni Unvan"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_rows"], 1)
        self.assertEqual(response.data["valid_rows"], 1)
        self.assertEqual(response.data["error_rows"], 0)
        self.assertNotIn("commit_rows", response.data)
        self.assertFalse(Employee.objects.filter(employee_code="IMP-001").exists())

    def test_employee_import_dry_run_csv_returns_preview_without_db_write(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_csv_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "E-posta", "Aktif Mi"],
                ["İdil Aksoy", "IMP-CSV-001", "idil.import@example.com", "Evet"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["format"], "csv")
        self.assertEqual(response.data["valid_rows"], 1)
        self.assertFalse(Employee.objects.filter(employee_code="IMP-CSV-001").exists())

    def test_employee_import_dry_run_requires_admin_role(self):
        self.client.force_authenticate(user=self.technician_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu"],
                ["Teknisyen Import", "IMP-TECH-001"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_employee_import_commit_blocks_error_rows(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu"],
                ["", "IMP-ERR-001"],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        commit_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )

        self.assertEqual(dry_run_response.status_code, status.HTTP_200_OK)
        self.assertEqual(dry_run_response.data["error_rows"], 1)
        self.assertEqual(commit_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(Employee.objects.filter(employee_code="IMP-ERR-001").exists())

    def test_employee_import_commit_creates_employee_master_data_and_audit_log(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "E-posta", "Departman", "Unvan", "Manager"],
                [
                    "Yeni Personel",
                    "IMP-OK-001",
                    "yeni.personel@example.com",
                    "Import Departmanı",
                    "Import Uzmanı",
                    "manager@example.com",
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(dry_run_response.status_code, status.HTTP_200_OK)

        with self.captureOnCommitCallbacks(execute=True):
            commit_response = self.client.post(
                "/api/employees/import/commit/",
                {
                    "import_id": dry_run_response.data["import_id"],
                    "mode": "create_only",
                },
                format="json",
            )

        self.assertEqual(commit_response.status_code, status.HTTP_200_OK)
        self.assertEqual(commit_response.data["created_count"], 1)

        imported_employee = Employee.objects.get(employee_code="IMP-OK-001")
        self.assertTrue(imported_employee.imported_from_excel)
        self.assertEqual(imported_employee.manager, self.manager)
        self.assertEqual(imported_employee.department.name, "Import Departmanı")
        self.assertEqual(imported_employee.job_title.name, "Import Uzmanı")

        audit_log = AuditLog.objects.filter(
            action=AuditLog.Action.CREATE,
            entity_type="employees.Import",
            metadata__operation="employee_import_commit",
        ).first()

        self.assertIsNotNone(audit_log)
        self.assertEqual(audit_log.actor, self.admin_user)
        metadata = audit_log.metadata
        self.assertEqual(metadata["module"], "employees")
        self.assertEqual(metadata["operation"], "employee_import_commit")
        self.assertEqual(metadata["import_id"], dry_run_response.data["import_id"])
        self.assertEqual(metadata["file_name"], "personnel-import.xlsx")
        self.assertEqual(metadata["format"], "xlsx")
        self.assertEqual(metadata["total_rows"], 1)
        self.assertEqual(metadata["created_count"], 1)
        self.assertEqual(metadata["skipped_count"], 0)
        self.assertEqual(metadata["error_count"], 0)
        self.assertEqual(metadata["warning_count"], 1)
        self.assertEqual(metadata["created_department_count"], 1)
        self.assertEqual(metadata["created_job_title_count"], 1)
        self.assertNotIn("rows", metadata)
        self.assertNotIn("row_data", metadata)
        self.assertNotIn("commit_rows", metadata)
        self.assertNotIn("emails", metadata)
        self.assertNotIn("phones", metadata)
        self.assertNotIn("yeni.personel@example.com", str(metadata))
        self.assertNotIn("manager@example.com", str(metadata))

    def test_employee_import_rejects_formula_cells(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "E-posta"],
                ["=HYPERLINK(\"https://example.com\")", "IMP-FORM-001", "formula@example.com"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 1)
        self.assertFalse(Employee.objects.filter(employee_code="IMP-FORM-001").exists())

    def test_employee_import_dry_run_supports_split_turkish_name_headers_xlsx(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                [
                    "Kullanıcı Adı",
                    "Adı",
                    "Soyadı",
                    "E-Posta adresi",
                    "Departman",
                    "Meslek",
                    "Durum",
                    "Test_Dosyasi_Durumu",
                ],
                [
                    "ahmet.yilmaz",
                    "Ahmet",
                    "Yılmaz",
                    "ahmet.yilmaz@example.com",
                    "İDARİ VE MALİ İŞLER MÜDÜRLÜĞÜ",
                    "BİLGİ İŞLEM",
                    "Aktif",
                    "OK",
                ],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["valid_rows"], 1)
        self.assertEqual(response.data["error_rows"], 0)
        self.assertIn("Test_Dosyasi_Durumu", response.data["unknown_headers"])

        normalized = response.data["rows"][0]["normalized"]
        self.assertEqual(normalized["full_name"], "Ahmet Yılmaz")
        self.assertEqual(normalized["email"], "ahmet.yilmaz@example.com")
        self.assertEqual(normalized["department_name"], "İDARİ VE MALİ İŞLER MÜDÜRLÜĞÜ")
        self.assertEqual(normalized["job_title_name"], "BİLGİ İŞLEM")
        self.assertTrue(normalized["is_active"])
        self.assertEqual(normalized["user_username"], "ahmet.yilmaz")
        self.assertFalse(Employee.objects.filter(email="ahmet.yilmaz@example.com").exists())

    def test_employee_import_commit_supports_split_turkish_name_headers_xlsx(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                [
                    "Kullanıcı Adı",
                    "Adı",
                    "Soyadı",
                    "E-Posta adresi",
                    "Departman",
                    "Meslek",
                    "Durum",
                    "Test_Dosyasi_Durumu",
                ],
                [
                    "ahmet.yilmaz",
                    "Ahmet",
                    "Yılmaz",
                    "ahmet.yilmaz@example.com",
                    "İDARİ VE MALİ İŞLER MÜDÜRLÜĞÜ",
                    "BİLGİ İŞLEM",
                    "Aktif",
                    "OK",
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(dry_run_response.status_code, status.HTTP_200_OK)

        with self.captureOnCommitCallbacks(execute=True):
            commit_response = self.client.post(
                "/api/employees/import/commit/",
                {
                    "import_id": dry_run_response.data["import_id"],
                    "mode": "create_only",
                },
                format="json",
            )

        self.assertEqual(commit_response.status_code, status.HTTP_200_OK)
        employee = Employee.objects.get(email="ahmet.yilmaz@example.com")
        self.assertEqual(employee.full_name, "Ahmet Yılmaz")
        self.assertEqual(employee.department.name, "İDARİ VE MALİ İŞLER MÜDÜRLÜĞÜ")
        self.assertEqual(employee.job_title.name, "BİLGİ İŞLEM")
        self.assertTrue(employee.is_active)
        self.assertIsNone(employee.user)

        self.assertTrue(
            AuditLog.objects.filter(
                action=AuditLog.Action.CREATE,
                entity_type="employees.Import",
                metadata__operation="employee_import_commit",
                metadata__import_id=dry_run_response.data["import_id"],
            ).exists(),
        )

    def test_employee_import_full_name_takes_precedence_over_split_name(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Adı", "Soyadı"],
                ["Ahmet Can Yılmaz", "Ahmet", "Yılmaz"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.data["rows"][0]["normalized"]["full_name"],
            "Ahmet Can Yılmaz",
        )

    def test_employee_import_missing_name_errors_when_no_full_name_or_split_name(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["E-posta", "Departman"],
                ["noname@example.com", "Import Departmanı"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 1)
        self.assertEqual(
            response.data["rows"][0]["errors"][0]["message"],
            "Ad Soyad zorunludur.",
        )

    def test_employee_import_unknown_headers_are_warnings_not_errors(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Test_Dosyasi_Durumu"],
                ["Bilinmeyen Kolon", "OK"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 0)
        self.assertIn("Test_Dosyasi_Durumu", response.data["unknown_headers"])

    def test_employee_import_status_alias_durum(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Durum"],
                ["Aktif Personel", "Aktif"],
                ["Pasif Personel", "Pasif"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["valid_rows"], 2)
        self.assertTrue(response.data["rows"][0]["normalized"]["is_active"])
        self.assertFalse(response.data["rows"][1]["normalized"]["is_active"])

    def test_employee_import_history_created_on_dry_run_without_pii(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "E-posta"],
                ["History Personel", "history.personel@example.com"],
            ],
            filename="history-import.xlsx",
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        job = EmployeeImportJob.objects.get(import_id=response.data["import_id"])
        self.assertEqual(job.status, EmployeeImportJob.Status.DRY_RUN)
        self.assertEqual(job.file_name, "history-import.xlsx")
        self.assertEqual(job.total_rows, 1)
        self.assertNotIn("history.personel@example.com", str(job.summary))
        self.assertFalse(hasattr(job, "rows"))

    def test_employee_import_history_updated_on_commit(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu"],
                ["History Commit", "IMP-HISTORY-001"],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        commit_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )

        self.assertEqual(commit_response.status_code, status.HTTP_200_OK)
        job = EmployeeImportJob.objects.get(import_id=dry_run_response.data["import_id"])
        self.assertEqual(job.status, EmployeeImportJob.Status.COMMITTED)
        self.assertEqual(job.created_count, 1)
        self.assertIsNotNone(job.committed_at)

    def test_employee_import_history_admin_only(self):
        self.client.force_authenticate(user=self.technician_user)

        response = self.client.get("/api/employees/import/history/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_employee_import_error_report_csv_returns_errors_and_warnings_without_pii(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "E-posta", "Telefon", "Departman"],
                ["", "secret.person@example.com", "5559998877", "Yeni Rapor Departmanı"],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        response = self.client.get(
            f"/api/employees/import/{dry_run_response.data['import_id']}/errors.csv/",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(codecs.BOM_UTF8))
        content = response.content.decode("utf-8-sig")
        self.assertIn("Satır;Durum;Alan;Mesaj;Değer", content)
        self.assertIn("Ad Soyad zorunludur.", content)
        self.assertIn("Commit sırasında yeni master data oluşturulacak.", content)
        self.assertNotIn("secret.person@example.com", content)
        self.assertNotIn("5559998877", content)

    def test_employee_import_dry_run_user_action_none_without_user_fields(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad"],
                ["User None Personel"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["rows"][0]["normalized"]["user_action"], "none")

    def test_employee_import_dry_run_user_action_link_existing_user(self):
        link_user = self.create_user_with_role("link-user", UserProfile.Role.REQUESTER)
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "User Username", "User Email", "User Role"],
                ["Link Personel", link_user.username, link_user.email, UserProfile.Role.REQUESTER],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        normalized = response.data["rows"][0]["normalized"]
        self.assertEqual(normalized["user_action"], "link_existing")
        self.assertEqual(normalized["user_id"], link_user.id)

    def test_employee_import_dry_run_user_conflict_username_email_different_users(self):
        username_user = self.create_user_with_role("conflict-username", UserProfile.Role.REQUESTER)
        email_user = self.create_user_with_role("conflict-email", UserProfile.Role.REQUESTER)
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "User Username", "User Email", "User Role"],
                ["Conflict Personel", username_user.username, email_user.email, UserProfile.Role.REQUESTER],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 1)
        self.assertEqual(response.data["rows"][0]["normalized"]["user_action"], "conflict")

    def test_employee_import_dry_run_user_existing_already_linked_to_other_employee(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "User Username", "User Email", "User Role"],
                [
                    "Already Linked",
                    self.requester_user.username,
                    self.requester_user.email,
                    UserProfile.Role.REQUESTER,
                ],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 1)
        self.assertEqual(response.data["rows"][0]["normalized"]["user_action"], "conflict")

    def test_employee_import_dry_run_new_user_requires_role(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "User Username", "User Email"],
                ["Needs Role", "needs.role", "needs.role@example.com"],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 1)
        self.assertEqual(response.data["rows"][0]["normalized"]["user_action"], "invalid")

    def test_employee_import_dry_run_blocks_admin_role_import(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "User Username", "User Email", "User Role"],
                ["Admin Import", "admin.import", "admin.import@example.com", UserProfile.Role.ADMIN],
            ],
        )

        response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_rows"], 1)
        self.assertEqual(response.data["rows"][0]["normalized"]["user_action"], "invalid")

    def test_employee_import_commit_links_existing_user_to_employee(self):
        link_user = self.create_user_with_role("commit-link-user", UserProfile.Role.REQUESTER)
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "User Username", "User Email", "User Role"],
                [
                    "Commit Link",
                    "IMP-LINK-001",
                    link_user.username,
                    link_user.email,
                    UserProfile.Role.REQUESTER,
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        commit_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )

        self.assertEqual(commit_response.status_code, status.HTTP_200_OK)
        employee = Employee.objects.get(employee_code="IMP-LINK-001")
        self.assertEqual(employee.user, link_user)
        self.assertEqual(commit_response.data["linked_user_count"], 1)

    def test_employee_import_commit_creates_inactive_user_with_unusable_password(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "User Username", "User Email", "User Role"],
                [
                    "New User Import",
                    "IMP-USER-001",
                    "new.import.user",
                    "new.import.user@example.com",
                    UserProfile.Role.REQUESTER,
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        commit_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )

        self.assertEqual(commit_response.status_code, status.HTTP_200_OK)
        user = User.objects.get(username="new.import.user")
        self.assertFalse(user.is_active)
        self.assertFalse(user.has_usable_password())
        self.assertEqual(user.profile.role, UserProfile.Role.REQUESTER)
        employee = Employee.objects.get(employee_code="IMP-USER-001")
        self.assertEqual(employee.user, user)
        self.assertEqual(commit_response.data["created_user_count"], 1)

    def test_employee_import_commit_blocks_when_user_conflict_errors(self):
        username_user = self.create_user_with_role("block-conflict-username", UserProfile.Role.REQUESTER)
        email_user = self.create_user_with_role("block-conflict-email", UserProfile.Role.REQUESTER)
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "User Username", "User Email", "User Role"],
                [
                    "Blocked Conflict",
                    "IMP-CONFLICT-001",
                    username_user.username,
                    email_user.email,
                    UserProfile.Role.REQUESTER,
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        commit_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )

        self.assertEqual(commit_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(Employee.objects.filter(employee_code="IMP-CONFLICT-001").exists())

    def test_employee_import_commit_audit_metadata_includes_user_counts_without_pii(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "User Username", "User Email", "User Role"],
                [
                    "Audit User Import",
                    "IMP-AUDIT-USER-001",
                    "audit.import.user",
                    "audit.import.user@example.com",
                    UserProfile.Role.REQUESTER,
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )

        with self.captureOnCommitCallbacks(execute=True):
            commit_response = self.client.post(
                "/api/employees/import/commit/",
                {
                    "import_id": dry_run_response.data["import_id"],
                    "mode": "create_only",
                },
                format="json",
            )

        self.assertEqual(commit_response.status_code, status.HTTP_200_OK)
        audit_log = AuditLog.objects.filter(
            action=AuditLog.Action.CREATE,
            entity_type="employees.Import",
            metadata__operation="employee_import_commit",
            metadata__import_id=dry_run_response.data["import_id"],
        ).first()

        self.assertIsNotNone(audit_log)
        metadata = audit_log.metadata
        self.assertEqual(metadata["created_user_count"], 1)
        self.assertEqual(metadata["linked_user_count"], 0)
        self.assertEqual(metadata["inactive_user_created_count"], 1)
        self.assertNotIn("audit.import.user@example.com", str(metadata))

    def test_employee_import_commit_idempotency_no_duplicate_users(self):
        self.client.force_authenticate(user=self.admin_user)
        upload = self.create_import_upload(
            [
                ["Ad Soyad", "Personel Kodu", "User Username", "User Email", "User Role"],
                [
                    "Idempotent User",
                    "IMP-IDEMP-001",
                    "idempotent.user",
                    "idempotent.user@example.com",
                    UserProfile.Role.REQUESTER,
                ],
            ],
        )

        dry_run_response = self.client.post(
            "/api/employees/import/dry-run/",
            {"file": upload},
            format="multipart",
        )
        first_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )
        second_response = self.client.post(
            "/api/employees/import/commit/",
            {
                "import_id": dry_run_response.data["import_id"],
                "mode": "create_only",
            },
            format="json",
        )

        self.assertEqual(first_response.status_code, status.HTTP_200_OK)
        self.assertEqual(second_response.status_code, status.HTTP_409_CONFLICT)
        self.assertEqual(User.objects.filter(username="idempotent.user").count(), 1)

    def test_repair_mojibake_for_export_handles_mixed_latin1_cp1252_turkish_text(self):
        self.assertEqual(
            repair_mojibake_for_export(REAL_WORLD_MOJIBAKE_DEPARTMENT),
            REPAIRED_REAL_WORLD_DEPARTMENT,
        )

    def test_repair_mojibake_for_export_keeps_valid_turkish_text_unchanged(self):
        values = [
            REPAIRED_REAL_WORLD_DEPARTMENT,
            "\u00c7a\u011fr\u0131 \u015eahin",
            REPAIRED_JOB_TITLE,
            "Normal English string",
        ]

        for value in values:
            with self.subTest(value=value):
                self.assertEqual(repair_mojibake_for_export(value), value)

    def test_employee_export_preserves_turkish_characters(self):
        self.department.name = REPAIRED_DEPARTMENT
        self.department.save(update_fields=["name"])
        self.job_title.name = REPAIRED_JOB_TITLE
        self.job_title.save(update_fields=["name"])
        self.employee.full_name = "\u00c7a\u011fr\u0131 \u015eahin"
        self.employee.save(update_fields=["full_name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(codecs.BOM_UTF8))

        content = response.content.decode("utf-8-sig")

        self.assertIn(REPAIRED_DEPARTMENT, content)
        self.assertIn(REPAIRED_JOB_TITLE, content)
        self.assertIn("\u00c7a\u011fr\u0131 \u015eahin", content)
        self.assertNotIn("\u00c3\u201e\u00c2\u00b0", content)
        self.assertNotIn("\u00c3\u2026", content)
        self.assertNotIn("\u00c3\u0192", content)

        employee_row = self.exported_employee_row(response, self.employee)

        self.assertEqual(employee_row[1], "\u00c7a\u011fr\u0131 \u015eahin")
        self.assertEqual(
            employee_row[6],
            REPAIRED_DEPARTMENT,
        )
        self.assertEqual(
            employee_row[7],
            REPAIRED_JOB_TITLE,
        )

    def test_export_uses_excel_compatible_semicolon_csv_with_utf8_bom(self):
        self.employee.full_name = "\u00c7a\u011fr\u0131; \u0130\u015fler"
        self.employee.phone = "\u0130lk sat\u0131r\n\u0130kinci sat\u0131r"
        self.employee.save(update_fields=["full_name", "phone"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(codecs.BOM_UTF8))

        content = response.content.decode("utf-8-sig")
        self.assertTrue(content.startswith("sep=;"))

        sep_line, rows = self.export_csv_rows(response)
        employee_row = self.exported_employee_row(response, self.employee)

        self.assertEqual(sep_line.rstrip("\r"), "sep=;")
        self.assertEqual(rows[0][1], "Ad Soyad")
        self.assertEqual(employee_row[1], "\u00c7a\u011fr\u0131; \u0130\u015fler")
        self.assertEqual(employee_row[4], "\u0130lk sat\u0131r\n\u0130kinci sat\u0131r")

    def test_employee_export_repairs_common_mojibake_if_source_value_is_corrupted(self):
        self.department.name = MOJIBAKE_DEPARTMENT
        self.department.save(update_fields=["name"])
        self.job_title.name = MOJIBAKE_JOB_TITLE
        self.job_title.save(update_fields=["name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        employee_row = self.exported_employee_row(response, self.employee)

        self.assertEqual(
            employee_row[6],
            REPAIRED_DEPARTMENT,
        )
        self.assertEqual(
            employee_row[7],
            REPAIRED_JOB_TITLE,
        )

        self.department.refresh_from_db()
        self.job_title.refresh_from_db()
        self.assertEqual(self.department.name, MOJIBAKE_DEPARTMENT)
        self.assertEqual(self.job_title.name, MOJIBAKE_JOB_TITLE)

    def test_employee_export_repairs_exact_real_world_mojibake_department(self):
        self.department.name = REAL_WORLD_MOJIBAKE_DEPARTMENT
        self.department.save(update_fields=["name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(codecs.BOM_UTF8))

        content = response.content.decode("utf-8-sig")

        self.assertIn(REPAIRED_REAL_WORLD_DEPARTMENT, content)
        self.assertNotIn("\u00c3\u201e\u00c2\u00b0", content)
        self.assertNotIn("\u00c3\u0192\u00c5\u201c", content)
        self.assertNotIn("\u00c3\u2026", content)
        self.assertNotIn("\u00c5\u201c", content)

    def test_employee_export_endpoint_repairs_exact_real_world_mojibake_from_response_bytes(self):
        self.department.name = REAL_WORLD_MOJIBAKE_DEPARTMENT
        self.department.save(update_fields=["name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(codecs.BOM_UTF8))

        decoded = response.content.decode("utf-8-sig")

        self.assertIn("sep=;", decoded)
        self.assertIn(REPAIRED_REAL_WORLD_DEPARTMENT, decoded)
        self.assertNotIn(REAL_WORLD_MOJIBAKE_DEPARTMENT, decoded)
        self.assertNotIn("\u00c3\u201e\u00c2\u00b0DAR\u00c3\u201e\u00c2\u00b0", decoded)

    def test_employee_export_response_content_can_be_written_to_disk_and_read_as_utf8_sig(self):
        self.department.name = REAL_WORLD_MOJIBAKE_DEPARTMENT
        self.department.save(update_fields=["name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        exported_bytes = bytes(response.content)

        self.assertTrue(exported_bytes.startswith(codecs.BOM_UTF8))

        decoded = exported_bytes.decode("utf-8-sig")

        self.assertIn(REPAIRED_REAL_WORLD_DEPARTMENT, decoded)
        self.assertNotIn(REAL_WORLD_MOJIBAKE_DEPARTMENT, decoded)

    def test_employee_export_applies_safe_csv_cell_to_nested_department_and_job_title(self):
        self.department.name = MOJIBAKE_DEPARTMENT
        self.department.save(update_fields=["name"])
        self.job_title.name = MOJIBAKE_JOB_TITLE
        self.job_title.save(update_fields=["name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        employee_row = self.exported_employee_row(response, self.employee)

        self.assertEqual(employee_row[6], REPAIRED_DEPARTMENT)
        self.assertEqual(employee_row[7], REPAIRED_JOB_TITLE)

    def test_csv_injection_protection_still_runs_after_mojibake_repair(self):
        self.employee.full_name = '=HYPERLINK("https://example.com")'
        self.employee.employee_code = "+EMP-001"
        self.employee.email = "@requester.example"
        self.employee.phone = "-5551112233"
        self.employee.save(
            update_fields=["full_name", "employee_code", "email", "phone"],
        )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        employee_row = self.exported_employee_row(response, self.employee)

        self.assertEqual(employee_row[1], '\'=HYPERLINK("https://example.com")')
        self.assertEqual(employee_row[2], "'+EMP-001")
        self.assertEqual(employee_row[3], "'@requester.example")
        self.assertEqual(employee_row[4], "'-5551112233")

    def test_employee_excel_export_returns_valid_workbook_with_expected_structure(self):
        self.department.name = REPAIRED_REAL_WORLD_DEPARTMENT
        self.department.save(update_fields=["name"])
        self.job_title.name = REPAIRED_JOB_TITLE
        self.job_title.save(update_fields=["name"])
        self.employee.full_name = "\u00c7a\u011fr\u0131 \u015eahin"
        self.employee.save(update_fields=["full_name"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export.xlsx/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("personnel-export.xlsx", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "no-store")
        self.assertTrue(response.content.startswith(b"PK"))

        workbook = self.export_xlsx_workbook(response)
        worksheet = workbook["Personel"]

        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertTrue(worksheet.auto_filter.ref)
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            [
                "ID",
                "Ad Soyad",
                "Personel Kodu",
                "E-posta",
                "Telefon",
                "Aktif Mi",
                "Departman",
                "Unvan",
                "Manager",
                "User Username",
                "User Email",
                "User Role",
                "Sync Source",
                "External HR ID",
                "Olu\u015fturulma",
                "G\u00fcncellenme",
            ],
        )
        self.assertTrue(worksheet["A1"].font.bold)

        employee_row = self.exported_employee_xlsx_row(response, self.employee)

        self.assertEqual(employee_row[1], "\u00c7a\u011fr\u0131 \u015eahin")
        self.assertEqual(employee_row[6], REPAIRED_REAL_WORLD_DEPARTMENT)
        self.assertEqual(employee_row[7], REPAIRED_JOB_TITLE)

    def test_employee_excel_export_repairs_real_world_mojibake_without_changing_valid_turkish(self):
        self.department.name = REAL_WORLD_MOJIBAKE_DEPARTMENT
        self.department.save(update_fields=["name"])
        self.job_title.name = REPAIRED_JOB_TITLE
        self.job_title.save(update_fields=["name"])
        self.employee.full_name = "\u00c7ALI\u015eANLAR"
        self.employee.email = "bilgi-islem@example.com"
        self.employee.save(update_fields=["full_name", "email"])

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export.xlsx/")
        employee_row = self.exported_employee_xlsx_row(response, self.employee)

        self.assertEqual(employee_row[1], "\u00c7ALI\u015eANLAR")
        self.assertEqual(employee_row[3], "bilgi-islem@example.com")
        self.assertEqual(employee_row[6], REPAIRED_REAL_WORLD_DEPARTMENT)
        self.assertEqual(employee_row[7], "\u0130dari ve Mali \u0130\u015fler M\u00fcd\u00fcr\u00fc")

    def test_employee_excel_export_prevents_formula_injection(self):
        self.employee.full_name = '=HYPERLINK("https://example.com")'
        self.employee.employee_code = "+EMP-001"
        self.employee.email = "@requester.example"
        self.employee.phone = "-5551112233"
        self.employee.save(
            update_fields=["full_name", "employee_code", "email", "phone"],
        )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/employees/export.xlsx/")
        employee_row = self.exported_employee_xlsx_row(response, self.employee)

        self.assertEqual(employee_row[1], '\'=HYPERLINK("https://example.com")')
        self.assertEqual(employee_row[2], "'+EMP-001")
        self.assertEqual(employee_row[3], "'@requester.example")
        self.assertEqual(employee_row[4], "'-5551112233")

    def test_employee_excel_export_applies_filters_and_ignores_pagination(self):
        for index in range(30):
            Employee.objects.create(
                full_name=f"Bulk Personel {index:02d}",
                email=f"bulk{index:02d}@example.com",
                department=self.department,
                job_title=self.job_title,
                is_active=True,
            )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get(
            "/api/employees/export.xlsx/",
            {
                "search": "Bulk Personel",
                "page_size": "5",
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        worksheet = self.export_xlsx_workbook(response)["Personel"]
        names = [row[1] for row in worksheet.iter_rows(min_row=2, values_only=True)]

        self.assertEqual(len(names), 30)
        self.assertIn("Bulk Personel 00", names)
        self.assertIn("Bulk Personel 29", names)
        self.assertNotIn("Requester Personel", names)

    def test_technician_can_export_employees_as_xlsx(self):
        self.client.force_authenticate(user=self.technician_user)

        response = self.client.get("/api/employees/export.xlsx/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_viewer_cannot_export_employees_as_xlsx(self):
        self.client.force_authenticate(user=self.viewer_user)

        response = self.client.get("/api/employees/export.xlsx/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_employee_excel_export_creates_audit_log(self):
        self.client.force_authenticate(user=self.admin_user)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.get(
                "/api/employees/export.xlsx/",
                {
                    "search": "Requester",
                },
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        audit_log = AuditLog.objects.filter(
            action=AuditLog.Action.EXPORT,
            entity_type="employees.Employee",
            metadata__operation="employee_export",
            metadata__format="xlsx",
        ).first()

        self.assertIsNotNone(audit_log)
        self.assertEqual(audit_log.actor, self.admin_user)
        self.assertEqual(audit_log.metadata["exported_count"], 1)
        self.assertEqual(audit_log.metadata["applied_filters"]["search"], "Requester")

    def test_technician_can_export_employees_as_csv(self):
        self.client.force_authenticate(user=self.technician_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_viewer_cannot_export_employees(self):
        self.client.force_authenticate(user=self.viewer_user)

        response = self.client.get("/api/employees/export/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_export_applies_filters_and_search(self):
        Employee.objects.create(
            full_name="Pasif Personel",
            email="pasif@example.com",
            department=self.department,
            job_title=self.job_title,
            is_active=False,
        )
        Employee.objects.create(
            full_name="Aktif Başka Personel",
            email="aktif@example.com",
            department=self.department,
            job_title=self.job_title,
            is_active=True,
        )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get(
            "/api/employees/export/",
            {
                "search": "Requester",
                "user_role": UserProfile.Role.REQUESTER,
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        content = response.content.decode("utf-8-sig")

        self.assertIn("Requester Personel", content)
        self.assertNotIn("Aktif Başka Personel", content)
        self.assertNotIn("Pasif Personel", content)

    def test_export_ignores_pagination_and_exports_filtered_all_rows(self):
        for index in range(30):
            Employee.objects.create(
                full_name=f"Bulk Personel {index:02d}",
                email=f"bulk{index:02d}@example.com",
                department=self.department,
                job_title=self.job_title,
                is_active=True,
            )

        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get(
            "/api/employees/export/",
            {
                "search": "Bulk Personel",
                "page_size": "5",
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        content = response.content.decode("utf-8-sig")
        lines = [line for line in content.splitlines() if line.strip()]

        self.assertEqual(len(lines), 32)
        self.assertIn("Bulk Personel 00", content)
        self.assertIn("Bulk Personel 29", content)

    def test_export_creates_audit_log(self):
        self.client.force_authenticate(user=self.admin_user)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.get(
                "/api/employees/export/",
                {
                    "search": "Requester",
                },
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        audit_log = AuditLog.objects.filter(
            action=AuditLog.Action.EXPORT,
            entity_type="employees.Employee",
            metadata__operation="employee_export",
        ).first()

        self.assertIsNotNone(audit_log)
        self.assertEqual(audit_log.actor, self.admin_user)
        self.assertEqual(audit_log.metadata["format"], "csv")
        self.assertEqual(audit_log.metadata["row_count"], 1)
        self.assertEqual(audit_log.metadata["filters"]["search"], "Requester")
