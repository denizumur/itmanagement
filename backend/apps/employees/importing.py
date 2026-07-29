import csv
import re
import unicodedata
from io import BytesIO, StringIO

from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from apps.accounts.models import UserProfile
from apps.employees.models import Department, Employee, JobTitle


FIELD_HEADERS = {
    "full_name": {"ad soyad", "adsoyad", "full name", "fullname", "name", "personel"},
    "employee_code": {"personel kodu", "personelkodu", "employee code", "employee_code"},
    "email": {"e-posta", "eposta", "email", "mail"},
    "phone": {"telefon", "phone"},
    "is_active": {"aktif mi", "aktifmi", "aktif", "is active", "is_active"},
    "department_name": {"departman", "department", "department name"},
    "job_title_name": {"unvan", "gorev", "görev", "job title", "job_title"},
    "manager": {"manager", "yonetici", "yönetici", "manager email", "manager username"},
    "user_username": {"user username", "username", "kullanici adi", "kullanıcı adı"},
    "user_email": {"user email", "kullanici email", "kullanıcı email"},
    "user_role": {"user role", "rol", "role"},
    "sync_source": {"sync source", "kaynak"},
    "external_hr_id": {"external hr id", "external_hr_id", "harici hr id"},
}

MAX_LENGTHS = {
    "full_name": 180,
    "employee_code": 50,
    "phone": 30,
    "external_hr_id": 120,
    "sync_source": 20,
}

ALLOWED_ROLES = {choice.value for choice in UserProfile.Role}
ALLOWED_SYNC_SOURCES = {choice.value for choice in Employee.SyncSource}
MOJIBAKE_MARKERS = ("Ã", "Ä", "Å", "Â", "�")


FIELD_HEADERS.update(
    {
        "full_name": FIELD_HEADERS["full_name"]
        | {
            "adı soyadı",
            "isim soyisim",
            "personel adı soyadı",
        },
        "first_name": {"adı", "ad", "isim", "first name", "firstname"},
        "last_name": {"soyadı", "soyad", "soyisim", "last name", "lastname"},
        "email": FIELD_HEADERS["email"]
        | {"e-posta adresi", "eposta adresi", "mail adresi"},
        "is_active": FIELD_HEADERS["is_active"] | {"durum", "aktif pasif", "status"},
        "department_name": FIELD_HEADERS["department_name"]
        | {"departmanı", "birim", "müdürlük", "departman adı"},
        "job_title_name": FIELD_HEADERS["job_title_name"]
        | {"ünvan", "meslek", "görev", "pozisyon"},
        "manager": FIELD_HEADERS["manager"] | {"yönetici"},
        "user_username": FIELD_HEADERS["user_username"] | {"kullanıcı adı"},
        "user_email": FIELD_HEADERS["user_email"] | {"kullanıcı email"},
    },
)


def normalize_header(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[\s_\-./]+", " ", text).strip()


HEADER_TO_FIELD = {
    normalize_header(header): field
    for field, headers in FIELD_HEADERS.items()
    for header in headers
}


def normalize_bool(value):
    text = str(value or "").strip().lower()
    if text in {"", "true", "1", "evet", "aktif", "yes"}:
        return True
    if text in {"false", "0", "hayir", "hayır", "pasif", "no"}:
        return False
    raise ValueError("Aktif Mi alanı Evet/Hayır, true/false veya 1/0 olmalı.")


def has_mojibake(value):
    text = str(value or "")
    return any(marker in text for marker in MOJIBAKE_MARKERS)


def display_cell_value(value):
    if value is None:
        return ""
    return str(value)


def validate_email_value(value, field_label, errors):
    if not value:
        return
    try:
        validate_email(value)
    except ValidationError:
        errors.append({"field": field_label, "message": "Geçerli bir e-posta girin."})


def parse_csv_upload(uploaded_file):
    raw = uploaded_file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV UTF-8 olarak okunamadı.") from exc
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return list(csv.reader(StringIO(text), delimiter=delimiter))


def parse_xlsx_upload(uploaded_file):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(uploaded_file.read()), data_only=False, read_only=True)
    worksheet = workbook.worksheets[0]
    rows = []
    formula_cells = set()
    for row_index, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
        values = []
        for column_index, cell in enumerate(row):
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formula_cells.add((row_index, column_index))
                values.append("")
            else:
                values.append(value)
        rows.append(values)
    return rows, formula_cells


def parse_employee_import(uploaded_file):
    file_name = uploaded_file.name
    extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    if extension not in {"csv", "xlsx"}:
        raise ValueError("Sadece .csv ve .xlsx dosyaları desteklenir.")

    formula_cells = set()
    if extension == "csv":
        raw_rows = parse_csv_upload(uploaded_file)
    else:
        raw_rows, formula_cells = parse_xlsx_upload(uploaded_file)

    raw_rows = [list(row) for row in raw_rows]
    if not raw_rows:
        raise ValueError("Dosya boş.")

    headers = [str(value or "").strip() for value in raw_rows[0]]
    mapped_fields = {}
    unknown_headers = []
    for index, header in enumerate(headers):
        if not header:
            continue
        field = HEADER_TO_FIELD.get(normalize_header(header))
        if field:
            mapped_fields[index] = field
        else:
            unknown_headers.append(header)

    rows = []
    seen = {"employee_code": {}, "email": {}, "external_hr_id": {}}
    existing = {
        "employee_code": set(Employee.objects.exclude(employee_code__isnull=True).exclude(employee_code="").values_list("employee_code", flat=True)),
        "email": set(Employee.objects.exclude(email="").values_list("email", flat=True)),
        "external_hr_id": set(Employee.objects.exclude(external_hr_id="").values_list("external_hr_id", flat=True)),
    }

    for row_number, raw_row in enumerate(raw_rows[1:], start=2):
        if not any(str(value or "").strip() for value in raw_row):
            continue
        normalized = {}
        errors = []
        warnings = []
        for column_index, field in mapped_fields.items():
            value = raw_row[column_index] if column_index < len(raw_row) else ""
            value = str(value or "").strip()
            if (row_number, column_index) in formula_cells:
                errors.append({"field": field, "message": "Formül hücresi import edilmez."})
                continue
            if has_mojibake(value):
                warnings.append({"field": field, "message": "Bozuk encoding olabilir; değer otomatik düzeltilmedi."})
            normalized[field] = value

        full_name = normalized.get("full_name", "").strip()
        first_name = normalized.get("first_name", "").strip()
        last_name = normalized.get("last_name", "").strip()
        split_full_name = f"{first_name} {last_name}".strip()

        if full_name:
            if split_full_name and normalize_header(full_name) != normalize_header(split_full_name):
                warnings.append(
                    {
                        "field": "full_name",
                        "message": "Ad Soyad kolonu ad/soyad kolonlarından farklı; Ad Soyad öncelikli kullanıldı.",
                    },
                )
            normalized["full_name"] = full_name
        elif split_full_name:
            normalized["full_name"] = split_full_name
            full_name = split_full_name
            if first_name and not last_name:
                warnings.append(
                    {
                        "field": "last_name",
                        "message": "Soyadı kolonu boş veya bulunamadı.",
                    },
                )

        if not full_name:
            errors.append({"field": "full_name", "message": "Ad Soyad zorunludur."})

        if normalized.get("user_username"):
            warnings.append(
                {
                    "field": "user_username",
                    "message": "Kullanıcı hesabı bağlama P7b kapsamındadır.",
                },
            )

        for field, max_length in MAX_LENGTHS.items():
            if len(normalized.get(field, "")) > max_length:
                errors.append({"field": field, "message": f"Maksimum {max_length} karakter olmalı."})

        validate_email_value(normalized.get("email", ""), "email", errors)
        validate_email_value(normalized.get("user_email", ""), "user_email", errors)

        if "is_active" in normalized:
            try:
                normalized["is_active"] = normalize_bool(normalized["is_active"])
            except ValueError as exc:
                errors.append({"field": "is_active", "message": str(exc)})
        else:
            normalized["is_active"] = True

        role = normalized.get("user_role", "")
        if role and role not in ALLOWED_ROLES:
            errors.append({"field": "user_role", "message": "Bilinmeyen rol."})

        sync_source = normalized.get("sync_source", "") or Employee.SyncSource.EXCEL
        if sync_source not in ALLOWED_SYNC_SOURCES:
            errors.append({"field": "sync_source", "message": "Bilinmeyen kaynak."})
        normalized["sync_source"] = sync_source

        for field in ("employee_code", "email", "external_hr_id"):
            value = normalized.get(field, "")
            if not value:
                continue
            if value in seen[field]:
                errors.append({"field": field, "message": f"Dosya içinde duplicate. İlk satır: {seen[field][value]}"})
            seen[field][value] = row_number
            if value in existing[field]:
                errors.append({"field": field, "message": "Mevcut personel kaydıyla çakışıyor."})

        manager_value = normalized.get("manager", "")
        if manager_value:
            manager = Employee.objects.filter(email=manager_value).first() or Employee.objects.filter(full_name=manager_value).first()
            if manager:
                normalized["manager_id"] = manager.id
            else:
                warnings.append({"field": "manager", "message": "Manager bulunamadı; boş bırakılacak."})

        for field, model in (("department_name", Department), ("job_title_name", JobTitle)):
            value = normalized.get(field, "")
            if value and not model.objects.filter(name=value).exists():
                warnings.append({"field": field, "message": "Commit sırasında yeni master data oluşturulacak."})

        status_value = "error" if errors else "warning" if warnings else "valid"
        rows.append({
            "row_number": row_number,
            "status": status_value,
            "data": {
                headers[index]: display_cell_value(
                    raw_row[index] if index < len(raw_row) else "",
                )
                for index in range(len(headers))
            },
            "normalized": normalized,
            "errors": errors,
            "warnings": warnings,
        })

    if not rows:
        raise ValueError("Import edilebilir satır bulunamadı.")

    error_rows = sum(1 for row in rows if row["errors"])
    warning_rows = sum(1 for row in rows if row["warnings"])
    valid_rows = sum(1 for row in rows if not row["errors"])
    return {
        "file_name": file_name,
        "format": extension,
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "error_rows": error_rows,
        "warning_rows": warning_rows,
        "headers": headers,
        "mapped_fields": {headers[index]: field for index, field in mapped_fields.items()},
        "unknown_headers": unknown_headers,
        "rows": rows[:100],
        "commit_rows": [row["normalized"] for row in rows if not row["errors"]],
        "summary": {"creates": valid_rows, "updates": 0, "skipped": error_rows},
    }
