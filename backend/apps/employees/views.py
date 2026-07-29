import csv
from uuid import uuid4
from io import BytesIO, StringIO

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import ListAPIView, RetrieveAPIView
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.models import UserProfile
from apps.accounts.permissions import IsAdminRole, IsTechnicianOrAdminRole, IsViewerOrAboveRole
from apps.audit.models import AuditLog
from apps.audit.services import create_audit_log
from apps.common.pagination import StandardResultsPagination
from apps.employees.filters import EmployeeFilterSet
from apps.employees.importing import parse_employee_import
from apps.employees.models import Department, Employee, EmployeeImportJob, JobTitle
from apps.employees.serializers import EmployeeDetailSerializer, EmployeeListSerializer


EMPLOYEE_SEARCH_FIELDS = [
    "full_name",
    "email",
    "employee_code",
    "phone",
    "external_hr_id",
    "department__name",
    "job_title__name",
    "manager__full_name",
    "user__username",
    "user__email",
]

EMPLOYEE_ORDERING_FIELDS = [
    "full_name",
    "email",
    "employee_code",
    "created_at",
    "updated_at",
    "is_active",
    "sync_source",
    "department__name",
    "job_title__name",
    "manager__full_name",
    "user__username",
]

EMPLOYEE_IMPORT_CACHE_TTL_SECONDS = 30 * 60
EMPLOYEE_IMPORT_CACHE_PREFIX = "employees_import:"
EMPLOYEE_IMPORT_ERROR_REPORT_FIELDS_TO_MASK = {"email", "user_email", "phone"}
User = get_user_model()


MOJIBAKE_MARKERS = (
    "\u00c3",
    "\u00c4",
    "\u00c5",
    "\u00c2",
    "\u00c3\u0192",
    "\u00c3\u201e",
    "\u00c3\u2026",
    "\u00c3\u201a",
    "\u00c5\u201c",
    "\u00c5\u00b8",
    "\x9c",
    "\x9e",
)


def employee_base_queryset():
    return Employee.objects.select_related(
        "user",
        "user__profile",
        "department",
        "job_title",
        "manager",
    )


def apply_default_active_filter(queryset, query_params):
    if "is_active" not in query_params:
        return queryset.filter(is_active=True)

    return queryset


def apply_employee_filterset(queryset, query_params):
    filterset = EmployeeFilterSet(data=query_params, queryset=queryset)

    if not filterset.is_valid():
        return None, filterset.errors

    return filterset.qs, None


def apply_employee_search(queryset, search_value):
    normalized_search = (search_value or "").strip()

    if not normalized_search:
        return queryset

    search_query = Q()

    for field_name in EMPLOYEE_SEARCH_FIELDS:
        search_query |= Q(**{f"{field_name}__icontains": normalized_search})

    return queryset.filter(search_query)


def apply_employee_ordering(queryset, ordering_value):
    raw_ordering = (ordering_value or "full_name").strip()

    if not raw_ordering:
        return queryset.order_by("full_name")

    ordering_fields = []

    for item in raw_ordering.split(","):
        item = item.strip()

        if not item:
            continue

        descending = item.startswith("-")
        field_name = item[1:] if descending else item

        if field_name not in EMPLOYEE_ORDERING_FIELDS:
            continue

        ordering_fields.append(f"-{field_name}" if descending else field_name)

    if not ordering_fields:
        ordering_fields = ["full_name"]

    return queryset.order_by(*ordering_fields)


def get_filtered_employee_queryset_for_export(request):
    queryset = employee_base_queryset()
    queryset = apply_default_active_filter(queryset, request.query_params)

    queryset, errors = apply_employee_filterset(queryset, request.query_params)

    if errors:
        return None, errors

    queryset = apply_employee_search(
        queryset,
        request.query_params.get("search"),
    )
    queryset = apply_employee_ordering(
        queryset,
        request.query_params.get("ordering"),
    )

    return queryset, None


def get_export_filters_snapshot(query_params):
    ignored_keys = {"page", "page_size"}

    return {
        key: value
        for key, value in query_params.items()
        if key not in ignored_keys and value not in ["", None]
    }


def user_role(employee):
    profile = getattr(getattr(employee, "user", None), "profile", None)

    if not profile:
        return ""

    return profile.role


def looks_like_mojibake(value):
    return any(marker in value for marker in MOJIBAKE_MARKERS)


def encode_mojibake_bytes(value):
    repaired_bytes = bytearray()

    for character in value:
        codepoint = ord(character)

        if codepoint <= 255:
            repaired_bytes.append(codepoint)
            continue

        try:
            repaired_bytes.extend(character.encode("cp1252"))
        except UnicodeError:
            return None

    return bytes(repaired_bytes)


def repair_mojibake_for_export(value):
    if not isinstance(value, str):
        return value

    if not looks_like_mojibake(value):
        return value

    repaired = value

    for _ in range(4):
        candidates = []

        for encoding in ("latin1", "cp1252"):
            try:
                candidates.append(repaired.encode(encoding).decode("utf-8"))
            except UnicodeError:
                continue

        mojibake_bytes = encode_mojibake_bytes(repaired)

        if mojibake_bytes is not None:
            try:
                candidates.append(mojibake_bytes.decode("utf-8"))
            except UnicodeError:
                pass

        current_marker_count = sum(
            repaired.count(marker) for marker in MOJIBAKE_MARKERS
        )

        for candidate in candidates:
            candidate_marker_count = sum(
                candidate.count(marker) for marker in MOJIBAKE_MARKERS
            )

            if candidate == repaired or candidate_marker_count >= current_marker_count:
                continue

            repaired = candidate
            break
        else:
            break

    return repaired


def safe_csv_cell(value):
    value = repair_mojibake_for_export(value)

    if not isinstance(value, str):
        return value

    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"

    return value


def write_safe_csv_row(writer, row):
    writer.writerow([safe_csv_cell(value) for value in row])


EMPLOYEE_EXPORT_HEADERS = [
    "ID",
    "Ad Soyad",
    "Personel Kodu",
    "E-posta",
    "Telefon",
    "Aktif Mi",
    "Departman",
    "Unvan",
    "Manager",
    "User Username",
    "User Email",
    "User Role",
    "Sync Source",
    "External HR ID",
    "Oluşturulma",
    "Güncellenme",
]


def employee_export_row(employee):
    return [
        employee.id,
        employee.full_name,
        employee.employee_code or "",
        employee.email or "",
        employee.phone or "",
        "Evet" if employee.is_active else "Hayır",
        employee.department.name if employee.department else "",
        employee.job_title.name if employee.job_title else "",
        employee.manager.full_name if employee.manager else "",
        employee.user.username if employee.user else "",
        employee.user.email if employee.user else "",
        user_role(employee),
        employee.sync_source,
        employee.external_hr_id,
        employee.created_at.replace(tzinfo=None) if employee.created_at else "",
        employee.updated_at.replace(tzinfo=None) if employee.updated_at else "",
    ]


def safe_excel_cell(value):
    value = repair_mojibake_for_export(value)

    if not isinstance(value, str):
        return value

    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"

    return value


def apply_export_response_headers(response, filename):
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Cache-Control"] = "no-store"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response


def create_employee_export_audit_log(request, *, export_format, row_count):
    create_audit_log(
        request=request,
        action=AuditLog.Action.EXPORT,
        entity_type="employees.Employee",
        entity_id="",
        entity_repr="Employee Export",
        metadata={
            "module": "employees",
            "operation": "employee_export",
            "format": export_format,
            "row_count": row_count,
            "exported_count": row_count,
            "filters": get_export_filters_snapshot(request.query_params),
            "applied_filters": get_export_filters_snapshot(request.query_params),
        },
    )


class EmployeeListAPIView(ListAPIView):
    """
    Legacy endpoint.

    Mevcut frontend akışlarını kırmamak için düz array response döndürmeye devam eder.
    Yeni tablo/pagination altyapısı için EmployeeTableListAPIView kullanılır.
    """

    serializer_class = EmployeeListSerializer
    permission_classes = [IsViewerOrAboveRole]

    def get_queryset(self):
        return employee_base_queryset().filter(is_active=True).order_by("full_name")


class EmployeeTableListAPIView(ListAPIView):
    """
    N7a table endpoint.

    Server-side pagination, search, filtering ve ordering destekler.
    """

    serializer_class = EmployeeListSerializer
    permission_classes = [IsViewerOrAboveRole]
    pagination_class = StandardResultsPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_class = EmployeeFilterSet

    search_fields = EMPLOYEE_SEARCH_FIELDS
    ordering_fields = EMPLOYEE_ORDERING_FIELDS
    ordering = ["full_name"]

    def get_queryset(self):
        queryset = employee_base_queryset().order_by("full_name")

        queryset = apply_default_active_filter(
            queryset,
            self.request.query_params,
        )

        return queryset


class EmployeeDetailAPIView(RetrieveAPIView):
    serializer_class = EmployeeDetailSerializer
    permission_classes = [IsViewerOrAboveRole]

    def get_queryset(self):
        return employee_base_queryset()


class EmployeeExportAPIView(APIView):
    permission_classes = [IsTechnicianOrAdminRole]

    def get(self, request):
        queryset, errors = get_filtered_employee_queryset_for_export(request)

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        row_count = queryset.count()
        today = timezone.localdate().isoformat()
        filename = f"personnel-export-{today}.csv"

        create_employee_export_audit_log(
            request,
            export_format="csv",
            row_count=row_count,
        )

        csv_buffer = StringIO(newline="")
        csv_buffer.write("sep=;\r\n")
        writer = csv.writer(csv_buffer, delimiter=";")

        write_safe_csv_row(writer, EMPLOYEE_EXPORT_HEADERS)

        for employee in queryset:
            row = employee_export_row(employee)
            row[14] = employee.created_at.isoformat() if employee.created_at else ""
            row[15] = employee.updated_at.isoformat() if employee.updated_at else ""
            write_safe_csv_row(writer, row)

        csv_bytes = csv_buffer.getvalue().encode("utf-8-sig")

        response = HttpResponse(
            csv_bytes,
            content_type="text/csv; charset=utf-8",
        )

        return apply_export_response_headers(response, filename)


class EmployeeExcelExportAPIView(APIView):
    permission_classes = [IsTechnicianOrAdminRole]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            return Response(
                {
                    "detail": (
                        "Excel export dependency is not available. "
                        "Please try again later."
                    ),
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        queryset, errors = get_filtered_employee_queryset_for_export(request)

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        row_count = queryset.count()
        filename = "personnel-export.xlsx"

        create_employee_export_audit_log(
            request,
            export_format="xlsx",
            row_count=row_count,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Personel"
        worksheet.append(EMPLOYEE_EXPORT_HEADERS)

        header_fill = PatternFill(fill_type="solid", fgColor="D9E5F2")
        header_font = Font(bold=True, color="111827")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for employee in queryset:
            worksheet.append(
                [safe_excel_cell(value) for value in employee_export_row(employee)]
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, header in enumerate(EMPLOYEE_EXPORT_HEADERS, start=1):
            letter = get_column_letter(column_index)
            max_length = len(header)

            for cell in worksheet[letter]:
                value = cell.value

                if value is None:
                    continue

                max_length = max(max_length, len(str(value)))

                if column_index in (15, 16) and cell.row > 1:
                    cell.number_format = "yyyy-mm-dd hh:mm"

            worksheet.column_dimensions[letter].width = min(
                max(max_length + 2, 10),
                45,
            )

        output = BytesIO()
        workbook.save(output)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        return apply_export_response_headers(response, filename)


def employee_import_job_response(job):
    actor = job.actor
    return {
        "id": job.id,
        "import_id": job.import_id,
        "file_name": job.file_name,
        "file_format": job.file_format,
        "status": job.status,
        "mode": job.mode,
        "actor": actor.username if actor else "",
        "total_rows": job.total_rows,
        "valid_rows": job.valid_rows,
        "error_rows": job.error_rows,
        "warning_rows": job.warning_rows,
        "created_count": job.created_count,
        "skipped_count": job.skipped_count,
        "created_department_count": job.created_department_count,
        "created_job_title_count": job.created_job_title_count,
        "created_user_count": job.created_user_count,
        "linked_user_count": job.linked_user_count,
        "file_size": job.file_size,
        "unknown_headers": job.unknown_headers,
        "summary": job.summary,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "committed_at": job.committed_at.isoformat() if job.committed_at else None,
        "expires_at": job.expires_at.isoformat() if job.expires_at else None,
    }


def get_cached_import(import_id):
    return cache.get(f"{EMPLOYEE_IMPORT_CACHE_PREFIX}{import_id}")


def masked_report_value(field, value):
    value = str(value or "")
    if field in EMPLOYEE_IMPORT_ERROR_REPORT_FIELDS_TO_MASK:
        return ""
    if len(value) > 80:
        return f"{value[:77]}..."
    return value


def resolve_import_user(row):
    action = row.get("user_action")
    if action == "link_existing":
        user_id = row.get("user_id")
        if user_id:
            return User.objects.get(id=user_id), False

        username = row.get("user_username", "")
        user_email = row.get("user_email", "")
        return (
            User.objects.filter(Q(username=username) | Q(email=user_email)).first(),
            False,
        )

    if action == "create_new":
        user = User(
            username=row["user_username"],
            email=row["user_email"],
            is_active=False,
        )
        user.set_unusable_password()
        user.save()
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.role = row["user_role"]
        profile.save(update_fields=["role"])
        return user, True

    return None, False


class EmployeeImportHistoryAPIView(APIView):
    permission_classes = [IsAdminRole]

    def get(self, request):
        limit = min(int(request.query_params.get("limit", 20)), 100)
        jobs = EmployeeImportJob.objects.select_related("actor").order_by("-created_at")[:limit]
        return Response([employee_import_job_response(job) for job in jobs])


class EmployeeImportHistoryDetailAPIView(APIView):
    permission_classes = [IsAdminRole]

    def get(self, request, pk):
        try:
            job = EmployeeImportJob.objects.select_related("actor").get(pk=pk)
        except EmployeeImportJob.DoesNotExist:
            return Response({"detail": "Import geçmişi bulunamadı."}, status=status.HTTP_404_NOT_FOUND)

        return Response(employee_import_job_response(job))


class EmployeeImportErrorReportAPIView(APIView):
    permission_classes = [IsAdminRole]

    def get(self, request, import_id):
        cached = get_cached_import(import_id)
        if not cached:
            return Response(
                {"detail": "Import sonucu bulunamadı veya süresi doldu."},
                status=status.HTTP_404_NOT_FOUND,
            )

        csv_buffer = StringIO(newline="")
        writer = csv.writer(csv_buffer, delimiter=";")
        writer.writerow(["Satır", "Durum", "Alan", "Mesaj", "Değer"])

        for row in cached.get("rows", []):
            normalized = row.get("normalized", {})
            for item_type, items in (("Hata", row.get("errors", [])), ("Uyarı", row.get("warnings", []))):
                for item in items:
                    field = item.get("field", "")
                    writer.writerow(
                        [
                            row.get("row_number", ""),
                            item_type,
                            field,
                            item.get("message", ""),
                            masked_report_value(field, normalized.get(field, "")),
                        ],
                    )

        csv_bytes = csv_buffer.getvalue().encode("utf-8-sig")
        response = HttpResponse(csv_bytes, content_type="text/csv; charset=utf-8")
        return apply_export_response_headers(
            response,
            f"employee-import-errors-{import_id}.csv",
        )


class EmployeeImportDryRunAPIView(APIView):
    permission_classes = [IsAdminRole]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response(
                {"detail": "Import dosyası zorunludur."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            parsed = parse_employee_import(uploaded_file)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        import_id = str(uuid4())
        parsed["import_id"] = import_id
        expires_at = timezone.now() + timezone.timedelta(
            seconds=EMPLOYEE_IMPORT_CACHE_TTL_SECONDS,
        )
        job = EmployeeImportJob.objects.create(
            import_id=import_id,
            file_name=parsed["file_name"],
            file_format=parsed["format"],
            status=EmployeeImportJob.Status.DRY_RUN,
            actor=request.user if request.user.is_authenticated else None,
            total_rows=parsed["total_rows"],
            valid_rows=parsed["valid_rows"],
            error_rows=parsed["error_rows"],
            warning_rows=parsed["warning_rows"],
            skipped_count=parsed["error_rows"],
            file_size=getattr(uploaded_file, "size", 0) or 0,
            unknown_headers=parsed.get("unknown_headers", []),
            summary=parsed.get("summary", {}),
            expires_at=expires_at,
        )
        parsed["job_id"] = job.id
        cache.set(
            f"{EMPLOYEE_IMPORT_CACHE_PREFIX}{import_id}",
            {**parsed, "committed": False},
            timeout=EMPLOYEE_IMPORT_CACHE_TTL_SECONDS,
        )

        response_data = {key: value for key, value in parsed.items() if key != "commit_rows"}
        return Response(response_data, status=status.HTTP_200_OK)


class EmployeeImportCommitAPIView(APIView):
    permission_classes = [IsAdminRole]

    def post(self, request):
        import_id = request.data.get("import_id")
        mode = request.data.get("mode", "create_only")

        if mode != "create_only":
            return Response(
                {"detail": "P7 aşamasında sadece create_only desteklenir."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cached = get_cached_import(import_id)
        if not cached:
            return Response(
                {"detail": "Import sonucu bulunamadı veya süresi doldu."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if cached.get("committed"):
            return Response(
                {"detail": "Bu import daha önce commit edildi."},
                status=status.HTTP_409_CONFLICT,
            )

        if cached.get("error_rows", 0) > 0:
            return Response(
                {"detail": "Hatalı satır varken import commit edilemez."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_departments = 0
        created_job_titles = 0
        created_users = 0
        linked_users = 0
        created_employees = []

        with transaction.atomic():
            for row in cached.get("commit_rows", []):
                employee_code = row.get("employee_code") or None
                email = row.get("email", "")
                external_hr_id = row.get("external_hr_id", "")

                duplicate_query = Q()
                if employee_code:
                    duplicate_query |= Q(employee_code=employee_code)
                if email:
                    duplicate_query |= Q(email=email)
                if external_hr_id:
                    duplicate_query |= Q(external_hr_id=external_hr_id)

                if duplicate_query and Employee.objects.filter(duplicate_query).exists():
                    transaction.set_rollback(True)
                    return Response(
                        {"detail": "Commit sırasında duplicate personel tespit edildi."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if row.get("user_action") == "create_new":
                    if User.objects.filter(
                        Q(username=row.get("user_username", ""))
                        | Q(email=row.get("user_email", ""))
                    ).exists():
                        transaction.set_rollback(True)
                        return Response(
                            {"detail": "Commit sırasında kullanıcı duplicate tespit edildi."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                linked_user, was_created = resolve_import_user(row)
                if linked_user and Employee.objects.filter(user=linked_user).exists():
                    transaction.set_rollback(True)
                    return Response(
                        {"detail": "Commit sırasında kullanıcı zaten bağlı görünüyor."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if was_created:
                    created_users += 1
                elif linked_user:
                    linked_users += 1

                department = None
                if row.get("department_name"):
                    department, created = Department.objects.get_or_create(
                        name=row["department_name"],
                    )
                    created_departments += int(created)

                job_title = None
                if row.get("job_title_name"):
                    job_title, created = JobTitle.objects.get_or_create(
                        name=row["job_title_name"],
                    )
                    created_job_titles += int(created)

                employee = Employee.objects.create(
                    user=linked_user,
                    full_name=row["full_name"],
                    employee_code=employee_code,
                    email=email,
                    phone=row.get("phone", ""),
                    is_active=row.get("is_active", True),
                    department=department,
                    job_title=job_title,
                    manager_id=row.get("manager_id"),
                    external_hr_id=external_hr_id,
                    sync_source=row.get("sync_source") or Employee.SyncSource.EXCEL,
                    imported_from_excel=True,
                    import_batch_id=import_id,
                )
                created_employees.append(employee)

            create_audit_log(
                request=request,
                action=AuditLog.Action.CREATE,
                entity_type="employees.Import",
                entity_id=import_id,
                entity_repr=f"Employee import {import_id}",
                metadata={
                    "module": "employees",
                    "operation": "employee_import_commit",
                    "import_id": import_id,
                    "file_name": cached.get("file_name"),
                    "format": cached.get("format"),
                    "total_rows": cached.get("total_rows"),
                    "created_count": len(created_employees),
                    "updated_count": 0,
                    "skipped_count": 0,
                    "error_count": cached.get("error_rows"),
                    "warning_count": cached.get("warning_rows"),
                    "created_department_count": created_departments,
                    "created_job_title_count": created_job_titles,
                    "created_user_count": created_users,
                    "linked_user_count": linked_users,
                    "user_conflict_count": cached.get("summary", {}).get("user_actions", {}).get("conflict", 0),
                    "inactive_user_created_count": created_users,
                    "admin_role_blocked_count": sum(
                        1
                        for row in cached.get("rows", [])
                        for error in row.get("errors", [])
                        if error.get("field") == "user_role"
                        and "admin" in error.get("message", "").lower()
                    ),
                },
            )

            EmployeeImportJob.objects.filter(import_id=import_id).update(
                status=EmployeeImportJob.Status.COMMITTED,
                committed_at=timezone.now(),
                created_count=len(created_employees),
                skipped_count=0,
                created_department_count=created_departments,
                created_job_title_count=created_job_titles,
                created_user_count=created_users,
                linked_user_count=linked_users,
                summary={
                    **cached.get("summary", {}),
                    "commit": {
                        "created_count": len(created_employees),
                        "created_user_count": created_users,
                        "linked_user_count": linked_users,
                    },
                },
            )

        cache.set(
            f"{EMPLOYEE_IMPORT_CACHE_PREFIX}{import_id}",
            {**cached, "committed": True},
            timeout=EMPLOYEE_IMPORT_CACHE_TTL_SECONDS,
        )

        return Response(
            {
                "import_id": import_id,
                "created_count": len(created_employees),
                "updated_count": 0,
                "skipped_count": 0,
                "error_count": 0,
                "warning_count": cached.get("warning_rows", 0),
                "created_department_count": created_departments,
                "created_job_title_count": created_job_titles,
                "created_user_count": created_users,
                "linked_user_count": linked_users,
            },
            status=status.HTTP_200_OK,
        )
